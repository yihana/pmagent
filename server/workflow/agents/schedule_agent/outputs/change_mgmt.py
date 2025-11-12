from __future__ import annotations
import json
from pathlib import Path
from typing import List, Dict, Any
import logging

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# ▶ OPTIONAL: Plotly HTML 시각화 지원
try:
    import plotly.graph_objects as go
    _PLOTLY_OK = True
except Exception:
    _PLOTLY_OK = False

logger = logging.getLogger("schedule.change")


class CPMEngine:
    """DAG 기반 CPM 계산 (Forward / Backward / Float / Critical Path)"""
    @staticmethod
    def _topo(nodes: List[str], edges: List[tuple]) -> List[str]:
        from collections import defaultdict, deque
        indeg = defaultdict(int)
        graph = {n: [] for n in nodes}
        for u, v in edges:
            graph.setdefault(u, [])
            graph.setdefault(v, [])
            graph[u].append(v); indeg[v] += 1
            indeg.setdefault(u, 0)
        q = deque([n for n in nodes if indeg[n] == 0])
        order = []
        while q:
            x = q.popleft()
            order.append(x)
            for nx in graph[x]:
                indeg[nx] -= 1
                if indeg[nx] == 0:
                    q.append(nx)
        if len(order) != len(nodes):
            logger.warning("[CPM] 순환 의심 - 입력 WBS 확인 필요")
        return order

    @staticmethod
    def build_dag_and_schedule(tasks: List[Dict[str, Any]]) -> Dict[str, Any]:
        # tasks: [{"id","name","duration","dependencies"/"predecessors"}]
        nodes = [t["id"] for t in tasks]
        duration = {t["id"]: int(t.get("duration", 1) or 1) for t in tasks}

        # predecessors 통일
        preds = {}
        edges = []
        for t in tasks:
            p = t.get("predecessors")
            if p is None:
                p = t.get("dependencies", [])
            if isinstance(p, str):
                p = [x.strip() for x in p.split(",") if x.strip()]
            preds[t["id"]] = list(p or [])
            for d in preds[t["id"]]:
                edges.append((d, t["id"]))

        order = CPMEngine._topo(nodes, edges)

        # Forward pass
        ES, EF = {}, {}
        for n in order:
            ES[n] = max([EF[p] for p in preds.get(n, [])], default=0)
            EF[n] = ES[n] + duration[n]

        # Backward pass
        proj_finish = max(EF.values()) if EF else 0
        succs = {n: [] for n in nodes}
        for u, v in edges:
            succs[u].append(v)
        LS, LF = {}, {}
        for n in reversed(order):
            LF[n] = min([LS[s] for s in succs.get(n, [])], default=proj_finish)
            LS[n] = LF[n] - duration[n]

        FLOAT = {n: LS[n] - ES[n] for n in nodes}
        critical_path = [n for n in nodes if FLOAT[n] == 0]

        return {
            "ES": ES, "EF": EF, "LS": LS, "LF": LF,
            "FLOAT": FLOAT, "critical_path": critical_path,
            "project_duration": proj_finish
        }

    @staticmethod
    def visualize_cpm_png(result: Dict[str, Any], output_png: Path):
        cp = result.get("critical_path", [])
        ES, EF = result.get("ES", {}), result.get("EF", {})
        plt.figure(figsize=(10, max(3, len(cp)*0.6)))
        plt.title("Critical Path")
        y = 0
        for n in cp:
            plt.plot([ES[n], EF[n]], [y, y], linewidth=6, color="#d32f2f")
            plt.text(ES[n], y+0.2, n)
            y += 1
        plt.xlabel("Days")
        output_png.parent.mkdir(parents=True, exist_ok=True)
        plt.tight_layout()
        plt.savefig(output_png)
        plt.close()
        logger.info(f"[CPM] PNG 저장: {output_png}")

    @staticmethod
    def visualize_cpm_html(tasks: List[Dict[str, Any]], result: Dict[str, Any], output_html: Path):
        """
        Plotly가 있으면 HTML 인터랙티브 간트/바 차트, 없으면 PNG로 대체.
        - Critical Path는 강조 색상
        """
        ES, EF = result["ES"], result["EF"]
        FLOAT = result["FLOAT"]
        cp_set = set(result["critical_path"])

        if _PLOTLY_OK:
            bars = []
            for t in tasks:
                tid = t["id"]
                es, ef = ES.get(tid, 0), EF.get(tid, 0)
                bars.append(dict(
                    Task=tid,
                    Start=es,
                    Finish=ef,
                    Float=FLOAT.get(tid, 0),
                    Critical=("Yes" if tid in cp_set else "No"),
                    Name=t.get("name") or tid
                ))
            # 수동 figure (express 없이)
            fig = go.Figure()
            for b in bars:
                fig.add_trace(
                    go.Bar(
                        x=[b["Finish"] - b["Start"]],
                        y=[b["Task"]],
                        base=b["Start"],
                        orientation="h",
                        name=b["Name"],
                        marker_color="#d32f2f" if b["Critical"] == "Yes" else "#90caf9",
                        hovertemplate=(
                            f"<b>{b['Name']}</b><br>"
                            f"Task: {b['Task']}<br>"
                            f"Start: {b['Start']}<br>"
                            f"Finish: {b['Finish']}<br>"
                            f"Float: {b['Float']}<br>"
                            f"Critical: {b['Critical']}<extra></extra>"
                        )
                    )
                )
            fig.update_layout(
                title="Critical Path (Interactive)",
                barmode="stack",
                xaxis_title="Days",
                yaxis_title="Tasks",
                showlegend=False,
                template="plotly_white",
                height=max(350, 24*len(bars))
            )
            output_html.parent.mkdir(parents=True, exist_ok=True)
            fig.write_html(str(output_html), include_plotlyjs="cdn")
            logger.info(f"[CPM] HTML 저장(Plotly): {output_html}")
            return

        # Plotly가 없을 경우 PNG만 생성하도록 유도
        CPMEngine.visualize_cpm_png(result, output_html.with_suffix(".png"))


class ChangeManagementGenerator:
    """변경관리표 + CPM 계산/시각화(HTML/PNG)"""
    @staticmethod
    def generate(
        project_id: str,
        output_path: Path,
        wbs_data: Dict[str, Any] | None = None,
        changes: List[Dict[str, Any]] | None = None
    ) -> Dict[str, Any]:
        changes = changes or []
        tasks = ChangeManagementGenerator._flatten_wbs(wbs_data) if wbs_data else []

        if tasks:
            cpm = CPMEngine.build_dag_and_schedule(tasks)
            # HTML(가능 시) + PNG 생성
            html_path = output_path.with_name(output_path.stem + "_critical_path.html")
            png_path = output_path.with_name(output_path.stem + "_critical_path.png")
            try:
                CPMEngine.visualize_cpm_html(tasks, cpm, html_path)
            except Exception as e:
                logger.warning(f"[CPM] HTML 시각화 실패: {e} → PNG만 생성")
            try:
                CPMEngine.visualize_cpm_png(cpm, png_path)
            except Exception as e:
                logger.warning(f"[CPM] PNG 시각화 실패: {e}")
        else:
            cpm = {
                "ES": {}, "EF": {}, "LS": {}, "LF": {}, "FLOAT": {},
                "critical_path": [], "project_duration": 0
            }
            html_path = None
            png_path = None

        # 변경요청에 일정영향/주공정영향 간단 표기
        for ch in changes:
            impact = ch.get("impact", "")
            ch["schedule_delta_days"] = ChangeManagementGenerator._parse_days(impact)
            ch["critical_path_impact"] = bool(cpm["critical_path"])

        # Excel 출력
        ChangeManagementGenerator._write_excel(output_path, changes, cpm)

        return {
            "excel": str(output_path),
            "critical_path_png": str(png_path) if png_path else None,
            "critical_path_html": str(html_path) if html_path else None,
            "project_duration_days": cpm["project_duration"],
            "critical_path": cpm["critical_path"]
        }

    @staticmethod
    def _flatten_wbs(wbs_json: Dict[str, Any]) -> List[Dict[str, Any]]:
        nodes = []
        def walk(node):
            nodes.append({
                "id": node.get("id") or node.get("name"),
                "name": node.get("name"),
                "duration": node.get("duration", 1),
                "predecessors": node.get("dependencies") or node.get("predecessors") or []
            })
            for c in node.get("children", []) or []:
                walk(c)
        for n in (wbs_json or {}).get("nodes", []):
            walk(n)
        return nodes

    @staticmethod
    def _parse_days(impact_text: str) -> int:
        import re
        m = re.search(r'([+-]?\d+)\s*일', impact_text or "")
        return int(m.group(1)) if m else 0

    @staticmethod
    def _write_excel(path: Path, changes: List[Dict[str, Any]], cpm: Dict[str, Any]):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "변경관리"

        headers = [
            "변경ID","제목","요청자","요청일","상태",
            "영향(요약)","일정영향(일)","주공정영향","승인자","승인일",
        ]
        ws.append(headers)
        bold = Font(bold=True)
        for c in ws[1]:
            c.font = bold
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", fgColor="DDDDDD")
        ws.auto_filter.ref = "A1:J1"

        for ch in changes:
            ws.append([
                ch.get("change_id") or ch.get("id"),
                ch.get("title"),
                ch.get("requester"),
                ch.get("requested_at"),
                ch.get("status") or ch.get("decision"),
                ch.get("impact"),
                ch.get("schedule_delta_days", 0),
                "Y" if ch.get("critical_path_impact") else "N",
                ch.get("approver"),
                ch.get("approval_date"),
            ])

        # CPM 요약 시트
        ws2 = wb.create_sheet("CPM")
        ws2.append(["작업ID","ES","EF","LS","LF","FLOAT","CRITICAL"])
        for tid, es in cpm["ES"].items():
            ws2.append([
                tid, es, cpm["EF"][tid], cpm["LS"][tid], cpm["LF"][tid],
                cpm["FLOAT"][tid], "Y" if cpm["FLOAT"][tid] == 0 else "N"
            ])
        ws2.append([])
        ws2.append(["Project Duration (days)", cpm["project_duration"]])
        for c in ws2[1]:
            c.font = bold

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
        logger.info(f"[CHANGE] 변경관리표 저장: {path}")