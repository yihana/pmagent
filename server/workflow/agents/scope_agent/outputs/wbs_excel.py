import openpyxl
from pathlib import Path
from typing import Dict, Any, List
from openpyxl.styles import Font, PatternFill, Alignment


class WBSExcelGenerator:
    """WBS Excel 산출물 생성기 (PMP 표준 양식)"""
    
    @staticmethod
    def generate(wbs_data: Dict[str, Any], output_path: Path) -> str:
        """
        WBS JSON 데이터를 Excel 파일로 변환
        
        Args:
            wbs_data: WBS 계층 구조 (JSON)
            output_path: 저장할 파일 경로
            
        Returns:
            str: 생성된 파일의 절대 경로
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WBS"
        
        # 헤더 생성 (Image 4 기준)
        headers = [
            "WBS Level 1", "WBS Level 2", "WBS Level 3", "WBS Level 4", 
            "WBS Level 5", "WBS Level 6",
            "산출물", "책임자", "착수(%)", "시작일자", "완료일자", 
            "소요일(MD)", "투입인력수(명)", "중소요일(MD)",
            "시작실적", "완료실적", "소요일실적", "투입인력실적", "중소요일실적",
            "완료상태", "실적 진도율", "실적 진도율 누적방안"
        ]
        ws.append(headers)
        
        # 헤더 스타일링
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # WBS 데이터 평탄화 및 입력
        def flatten_wbs(node, level=1, parent_name=""):
            rows = []
            node_id = node.get("id", node.get("name"))
            node_name = node.get("name", node_id)
            
            # Level별 배치
            row = [""] * len(headers)
            row[level-1] = node_name  # Level 위치에 이름 배치
            row[6] = node.get("deliverables", "")  # 산출물
            row[7] = node.get("owner", "PM")       # 책임자
            row[8] = f"{node.get('progress', 0)}%" # 착수(%)
            row[9] = node.get("start_date", "")    # 시작일자
            row[10] = node.get("end_date", "")     # 완료일자
            row[11] = node.get("duration_md", "")  # 소요일(MD)
            row[12] = node.get("headcount", "1")   # 투입인력수
            row[13] = ""  # 중소요일(MD) - 계산 필요
            # 실적 필드는 비워둠 (14~18)
            row[19] = node.get("status", "")       # 완료상태
            row[20] = f"{node.get('completion', 0)}%"  # 실적 진도율
            row[21] = ""  # 누적방안
            
            rows.append(row)
            
            # 자식 노드 재귀 처리
            for child in node.get("children", []):
                rows.extend(flatten_wbs(child, level + 1, node_name))
            
            return rows
        
        # 루트 노드부터 변환
        all_rows = []
        for root in wbs_data.get("nodes", []):
            all_rows.extend(flatten_wbs(root))
        
        for row in all_rows:
            ws.append(row)
        
        # 열 너비 자동 조정
        column_widths = {
            'A': 20, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 20,  # Level 1-6
            'G': 25,  # 산출물
            'H': 12,  # 책임자
            'I': 10,  # 착수(%)
            'J': 12,  # 시작일자
            'K': 12,  # 완료일자
            'L': 12,  # 소요일(MD)
            'M': 14,  # 투입인력수
            'N': 14,  # 중소요일
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 행 높이 조정
        ws.row_dimensions[1].height = 30
        
        # 파일 저장
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return str(output_path.resolve())
    