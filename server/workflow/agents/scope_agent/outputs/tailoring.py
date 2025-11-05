# server/workflow/agents/scope_agent/output/
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment


class TailoringGenerator:
    """테일러링 Excel 생성기 (방법론 맞춤)"""
    
    @staticmethod
    def generate(methodology: str, output_path: Path) -> str:
        """
        테일러링 문서 생성
        
        Args:
            methodology: 프로젝트 방법론 (waterfall/agile)
            output_path: 저장할 파일 경로
            
        Returns:
            str: 생성된 파일의 절대 경로
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "테일러링"
        
        # 헤더
        headers = [
            "조직방법론분야", "단계", "활동", "태스크", "산출물", 
            "담당자", "템플릿명", "도구", "비고"
        ]
        ws.append(headers)
        
        # 헤더 스타일링
        header_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="000000", bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 방법론별 태스크 정의
        if methodology.lower() == "waterfall":
            tasks = TailoringGenerator._get_waterfall_tasks()
        else:  # agile
            tasks = TailoringGenerator._get_agile_tasks()
        
        # 데이터 입력
        for task in tasks:
            ws.append(task)
        
        # 열 너비 조정
        column_widths = {
            'A': 15,  # 조직방법론분야
            'B': 12,  # 단계
            'C': 20,  # 활동
            'D': 25,  # 태스크
            'E': 20,  # 산출물
            'F': 12,  # 담당자
            'G': 20,  # 템플릿명
            'H': 15,  # 도구
            'I': 15,  # 비고
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 모든 셀 줄바꿈 활성화
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # 파일 저장
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return str(output_path.resolve())
    
    @staticmethod
    def _get_waterfall_tasks():
        """Waterfall 방법론 태스크"""
        return [
            ["프로젝트 관리", "착수", "착수준비", "프로젝트 헌장 작성", "프로젝트 헌장", "PM", "", "MS Office", ""],
            ["프로젝트 관리", "착수", "착수준비", "범위 기술서 작성", "범위 기술서", "PM", "", "MS Office", ""],
            ["프로젝트 개발", "분석", "요구사항 정의", "인터뷰", "요구사항정의서", "BA, PM", "", "MS Office", ""],
            ["프로젝트 개발", "분석", "요구사항 정의", "워크샵", "기능정의서", "BA", "", "MS Office", ""],
            ["프로젝트 개발", "설계", "시스템 설계", "아키텍처 설계", "시스템 설계서", "Architect", "", "MS Office", ""],
            ["프로젝트 개발", "설계", "DB 설계", "ERD 작성", "DB 설계서", "DA", "", "ERWin", ""],
            ["프로젝트 개발", "설계", "UI/UX 설계", "화면설계", "화면설계서", "Designer", "", "Figma", ""],
            ["프로젝트 개발", "개발", "코딩", "프로그램 개발", "소스코드", "Developer", "", "IDE", ""],
            ["프로젝트 개발", "개발", "단위테스트", "단위테스트 수행", "테스트결과서", "Developer", "", "JUnit", ""],
            ["프로젝트 개발", "테스트", "통합테스트", "통합테스트 수행", "통합테스트결과서", "QA", "", "Selenium", ""],
            ["프로젝트 개발", "테스트", "UAT", "사용자 인수테스트", "UAT 결과서", "사용자", "", "", ""],
            ["프로젝트 개발", "이행", "배포", "운영환경 배포", "배포결과서", "PMO", "", "Jenkins", ""],
            ["프로젝트 개발", "종료", "프로젝트 종료", "완료보고서 작성", "완료보고서", "PM", "", "MS Office", ""],
        ]
    
    @staticmethod
    def _get_agile_tasks():
        """Agile 방법론 태스크"""
        return [
            ["프로젝트 관리", "Sprint 0", "백로그 작성", "User Story 작성", "Product Backlog", "PO", "", "Jira", ""],
            ["프로젝트 관리", "Sprint 0", "Sprint 계획", "Sprint Planning", "Sprint Backlog", "Scrum Master", "", "Jira", ""],
            ["프로젝트 개발", "Sprint 1", "개발", "스토리 개발", "Working Software", "Team", "", "IDE", ""],
            ["프로젝트 개발", "Sprint 1", "테스트", "스프린트 테스트", "테스트 결과", "QA", "", "Jest", ""],
            ["프로젝트 관리", "Sprint 1", "리뷰", "Sprint Review", "Sprint Review 회의록", "Team", "", "Zoom", ""],
            ["프로젝트 관리", "Sprint 1", "회고", "Retrospective", "개선사항", "Team", "", "Miro", ""],
            ["프로젝트 개발", "Sprint 2", "개발", "스토리 개발", "Working Software", "Team", "", "IDE", ""],
            ["프로젝트 개발", "Sprint 2", "테스트", "스프린트 테스트", "테스트 결과", "QA", "", "Jest", ""],
            ["프로젝트 관리", "Sprint 2", "리뷰", "Sprint Review", "Sprint Review 회의록", "Team", "", "Zoom", ""],
            ["프로젝트 관리", "Sprint 2", "회고", "Retrospective", "개선사항", "Team", "", "Miro", ""],
            ["프로젝트 관리", "매일", "Daily Standup", "Daily Scrum", "진행상황 공유", "Team", "", "Slack", ""],
            ["프로젝트 관리", "지속적", "백로그 관리", "Backlog Refinement", "정제된 Backlog", "PO", "", "Jira", ""],
        ]