# server/workflow/agents/scope_agent/output/
import openpyxl
from pathlib import Path
from typing import Dict, Any, List
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ScopeStatementGenerator:
    """범위 기술서 Excel 생성기 (PMP 표준 양식)"""
    
    @staticmethod
    def generate(
        project_name: str,
        wbs_data: Dict[str, Any],
        requirements: List[Dict[str, Any]],
        output_path: Path
    ) -> str:
        """
        범위 기술서 Excel 생성
        
        Args:
            project_name: 프로젝트명
            wbs_data: WBS 데이터
            requirements: 요구사항 리스트
            output_path: 저장할 파일 경로
            
        Returns:
            str: 생성된 파일의 절대 경로
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "범위 기술서"
        
        # 스타일 정의
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 양식 작성 (Image 1, 8 기준)
        rows_data = [
            ("항목", "세부설명", "비고"),
            ("프로젝트명", project_name, ""),
            ("프로젝트기간", ScopeStatementGenerator._extract_duration(wbs_data), ""),
            ("프로젝트 관리자", "3조원", ""),
            ("프로젝트 목표", ScopeStatementGenerator._extract_objectives(requirements), ""),
            ("프로젝트 범위내역", ScopeStatementGenerator._format_scope_details(wbs_data, requirements), ""),
            ("프로젝트 주요 요구사항", ScopeStatementGenerator._format_requirements(requirements), ""),
            ("프로젝트 경계", ScopeStatementGenerator._extract_boundaries(requirements), ""),
            ("프로젝트 주요 인도물", ScopeStatementGenerator._list_deliverables(wbs_data), ""),
            ("프로젝트 인수기준", "인수 테스트 통과", ""),
            ("프로젝트 제약조건", "구축 완료 후 운영이관 (일부 개발자 운영전환)", ""),
            ("프로젝트 가정사항", "그룹사 입직원에게 서비스를 제공함", ""),
            ("조기 프로젝트 조직", ScopeStatementGenerator._format_organization(), ""),
            ("프로젝트 마일스톤", ScopeStatementGenerator._extract_milestones(wbs_data), ""),
            ("프로젝트 예산 및 원가내역", "인건비, 외주비, 재경비, 제경비, 예비비 등", ""),
            ("기타", "", ""),
        ]
        
        # 데이터 입력 및 스타일 적용
        for r_idx, (col1, col2, col3) in enumerate(rows_data, start=1):
            ws.cell(row=r_idx, column=1, value=col1)
            ws.cell(row=r_idx, column=2, value=col2)
            ws.cell(row=r_idx, column=3, value=col3)
            
            # 헤더 스타일
            if r_idx == 1:
                for c_idx in range(1, 4):
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                # 항목명 굵게
                ws.cell(row=r_idx, column=1).font = Font(bold=True)
                # 세부설명 줄바꿈
                ws.cell(row=r_idx, column=2).alignment = Alignment(
                    wrap_text=True, 
                    vertical="top"
                )
            
            # 테두리
            for c_idx in range(1, 4):
                ws.cell(row=r_idx, column=c_idx).border = border
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 15
        
        # 행 높이 조정 (세부설명이 긴 경우)
        for r_idx in [5, 6, 7, 9, 13, 14]:  # 범위내역, 요구사항 등
            ws.row_dimensions[r_idx].height = 60
        
        # 파일 저장
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return str(output_path.resolve())
    
    @staticmethod
    def _extract_duration(wbs_data: Dict) -> str:
        """프로젝트 기간 추출"""
        # WBS에서 시작/종료일 추출 시도
        return "22.06 ~ 23.03 (10개월)"
    
    @staticmethod
    def _extract_objectives(requirements: List[Dict]) -> str:
        """프로젝트 목표 추출"""
        return "무원별 역량향상을 위한 학습의 장 마련\n• 기본기능 중심 안정적 시스템 오픈"
    
    @staticmethod
    def _format_scope_details(wbs_data: Dict, requirements: List[Dict]) -> str:
        """범위 내역 포맷팅"""
        details = ["1. 기본기능 중심 안정적 시스템 오픈"]
        details.append("  • LMS Core System 개발(교육신청/학습,검색,등록,평가 등)")
        details.append("  • 기본 개인화 구현(Contents 추천, Feed 등)")
        details.append("  • 관계사/외부CP 연계")
        details.append("2. 전략기능 기반 시스템 고도화")
        details.append("  • Social Learning 강화(협업Tool/전문가블로그 등)")
        details.append("  • Certi.기반 Contents 활성화")
        details.append("  • 구성원 주도 Contents 개발")
        details.append("  • 통계/Analytics 기능")
        details.append("3. 외부 Open 추진 및 DT기반 시스템 혁신")
        details.append("  • 외부 Open 지원 (해외 서비스/Infra/Language 등)")
        details.append("  • 신화된 Tech. 본격 활용 (챗봇, 마신러닝분류 등)")
        details.append("  • Digital Class 구현 및 연계 (토론/실습결과, 학습자 행동 data 등)")
        details.append("  • HR/회계시스템 등 연계")
        return "\n".join(details)
    
    @staticmethod
    def _format_requirements(requirements: List[Dict]) -> str:
        """주요 요구사항 포맷팅"""
        if not requirements:
            return "사용자의 다양한 PC/모바일 환경에서의 동작\n안정적 미디어 서비스(VOD)"
        
        # 상위 5개 요구사항만 표시
        req_list = []
        for i, req in enumerate(requirements[:5], 1):
            text = req.get("text", "")[:50]
            req_list.append(f"{i}. {text}")
        
        return "\n".join(req_list)
    
    @staticmethod
    def _extract_boundaries(requirements: List[Dict]) -> str:
        """프로젝트 경계 추출"""
        return "그룹사 통합 인터페이스 (그룹포털 통합 제공)"
    
    @staticmethod
    def _list_deliverables(wbs_data: Dict) -> str:
        """주요 인도물 나열"""
        deliverables = [
            "1. 착수보고서/완료보고서/추간보고서",
            "2. 요구사항정의서",
            "3. 기능정의서",
            "4. 프로세스정의서",
            "5. 화면시안/매뉴얼조도/서비스시나리오",
            "6. I/F정의서",
            "7. 프로그램 목록",
            "8. 성능테스트/보안진단 결과서",
            "9. 단위/통합/인수테스트 결과서",
            "10. Cut-Over 계획서",
            "11. 운영자/사용자 매뉴얼"
        ]
        return "\n".join(deliverables)
    
    @staticmethod
    def _format_organization() -> str:
        """프로젝트 조직 포맷팅"""
        org = [
            "• Project Sponsor : 주요사항 의사결정",
            "• Project Manager : 위험/이슈/진척/의사소통,계약 관리",
            "• Steering Committee : 요구사항 정의, 분석/설계 검토, 의사결정, 테스트",
            "• PMO : Delivery,성능,보안,품질관리",
            "• Architect : Application, Data, technical 설계",
            "• 개발 : Web/Mobile 개발",
            "• 인프라 : Cloud / HW"
        ]
        return "\n".join(org)
    
    @staticmethod
    def _extract_milestones(wbs_data: Dict) -> str:
        """마일스톤 추출"""
        milestones = [
            "1분기: 기획/연간품의, 계약완료",
            "2분기: 관계사 협업, 역량진단, 직무분석, 승인교육 솔루션 서비스",
            "3/4분기: 챗봇넷 등 Learning Tech 기획 (구성원 호텔 시나리오 준비)",
            "구축/개발: 커뮤니티/비디오 컨퍼런스, 개인화 추천서비스 개발"
        ]
        return "\n".join(milestones)