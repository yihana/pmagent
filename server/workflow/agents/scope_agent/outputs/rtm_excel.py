import openpyxl
from pathlib import Path
from typing import List, Dict, Any
from openpyxl.styles import Font, PatternFill, Alignment


class RTMExcelGenerator:
    """요구사항 추적표(RTM) Excel 생성기"""
    
    @staticmethod
    def generate(requirements: List[Dict[str, Any]], output_path: Path) -> str:
        """
        요구사항 데이터를 RTM Excel로 변환
        
        Args:
            requirements: 요구사항 리스트
            output_path: 저장할 파일 경로
            
        Returns:
            str: 생성된 파일의 절대 경로
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "요구사항 추적표"
        
        # 헤더 (Image 5 기준)
        headers = [
            "번호", "분야", "요구사항ID", "구분", "출처", "요구사항ID2", "요구사항명",
            "[설계단계] 확정명(ID)", "[설계단계] 관련 DB Entity", 
            "[설계단계] 연계 인터페이스ID",
            "[개발단계] 프로그램명(ID)", 
            "[테스트단계] 단위테스트명(ID)", 
            "[테스트단계] 통합테스트명(ID)",
            "비고"
        ]
        ws.append(headers)
        
        # 헤더 스타일링
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="000000", bold=True, size=10)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # RTM 데이터 입력
        for idx, req in enumerate(requirements, start=1):
            req_id = req.get("id", f"SFR-{idx:03d}")
            category = req.get("category", "web")
            req_type = req.get("type", "기능")
            source = req.get("source", "RFP")
            req_text = req.get("text", req.get("requirement", ""))
            
            row = [
                idx,                              # 번호
                category,                         # 분야
                req_id,                          # 요구사항ID
                req_type,                        # 구분
                source,                          # 출처
                f"{req_id}-001",                 # 요구사항ID2
                req_text[:50] + "..." if len(req_text) > 50 else req_text,  # 요구사항명
                f"설계_{req_id}",                # 확정명
                "없음",                          # DB Entity
                "없음",                          # 인터페이스ID
                f"프로그램_{idx:03d}",          # 프로그램명
                f"단위테스트_{idx:03d}",        # 단위테스트
                f"통합테스트_{idx:03d}",        # 통합테스트
                ""                               # 비고
            ]
            ws.append(row)
        
        # 열 너비 조정
        column_widths = {
            'A': 8,   # 번호
            'B': 12,  # 분야
            'C': 15,  # 요구사항ID
            'D': 10,  # 구분
            'E': 10,  # 출처
            'F': 15,  # 요구사항ID2
            'G': 40,  # 요구사항명
            'H': 20,  # 확정명
            'I': 20,  # DB Entity
            'J': 20,  # 인터페이스ID
            'K': 20,  # 프로그램명
            'L': 25,  # 단위테스트
            'M': 25,  # 통합테스트
            'N': 15,  # 비고
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 행 높이 조정
        ws.row_dimensions[1].height = 30
        
        # 모든 셀 텍스트 줄바꿈 활성화
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # 파일 저장
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        return str(output_path.resolve())